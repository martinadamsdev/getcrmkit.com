from __future__ import annotations

import tempfile
import uuid
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from app.domain.customer.entities import Contact, Customer
from app.domain.customer.enums import DataJobStatus
from app.domain.customer.services import DuplicateChecker
from app.domain.customer.value_objects import CustomerFilter
from app.infra.database.connection import async_session_factory
from app.infra.database.repositories.customer_repository import (
    ContactRepository,
    CustomerGradeRepository,
    CustomerRepository,
)
from app.infra.database.repositories.data_job_repository import DataJobRepository

BATCH_SIZE = 100

EXPORT_HEADERS = [
    "客户名",
    "国家",
    "地区",
    "城市",
    "行业",
    "公司规模",
    "网站",
    "来源",
    "等级",
    "跟进阶段",
    "主营产品",
    "年采购量",
    "现有供应商",
    "联系人姓名",
    "联系人邮箱",
    "联系人电话",
    "WhatsApp",
]

_FIELD_MAP: dict[str, str] = {
    "name": "name",
    "客户名": "name",
    "country": "country",
    "国家": "country",
    "region": "region",
    "地区": "region",
    "city": "city",
    "城市": "city",
    "industry": "industry",
    "行业": "industry",
    "company_size": "company_size",
    "公司规模": "company_size",
    "website": "website",
    "网站": "website",
    "source": "source",
    "来源": "source",
    "main_products": "main_products",
    "主营产品": "main_products",
    "annual_volume": "annual_volume",
    "年采购量": "annual_volume",
    "current_supplier": "current_supplier",
    "现有供应商": "current_supplier",
    "contact_name": "contact_name",
    "联系人姓名": "contact_name",
    "contact_email": "contact_email",
    "联系人邮箱": "contact_email",
    "contact_phone": "contact_phone",
    "联系人电话": "contact_phone",
    "whatsapp": "whatsapp",
}


async def process_customer_import(ctx: dict[str, Any], *, job_id: str) -> None:
    """Process customer import from Excel file."""
    async with async_session_factory() as session:
        job_repo = DataJobRepository(session)
        customer_repo = CustomerRepository(session)
        contact_repo = ContactRepository(session)
        duplicate_checker = DuplicateChecker(customer_repo=customer_repo)

        job = await job_repo.get_by_job_id(uuid.UUID(job_id))
        if job is None:
            return

        job.status = DataJobStatus.PROCESSING
        job.started_at = datetime.now(UTC)
        await job_repo.update(job)
        await session.commit()

        errors: list[dict[str, Any]] = []
        success_count = 0
        processed = 0

        try:
            wb = load_workbook(job.file_path, read_only=True)
            ws = wb.active
            header_row = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            col_indices: dict[str, int] = {}
            for i, h in enumerate(header_row):
                if h in _FIELD_MAP:
                    col_indices[_FIELD_MAP[h]] = i

            job.total_rows = len(rows)
            await job_repo.update(job)
            await session.commit()

            for batch_start in range(0, len(rows), BATCH_SIZE):
                batch = rows[batch_start : batch_start + BATCH_SIZE]
                for row_idx, row in enumerate(batch, start=batch_start + 2):
                    try:

                        def get_val(field: str, _row: Any = row) -> str | None:
                            idx = col_indices.get(field)
                            if idx is not None and idx < len(_row) and _row[idx]:
                                return str(_row[idx]).strip()
                            return None

                        name = get_val("name")
                        if not name:
                            errors.append({"row": row_idx, "error": "name is required"})
                            processed += 1
                            continue

                        email = get_val("contact_email")
                        dupes = await duplicate_checker.check(tenant_id=job.tenant_id, name=name, email=email)
                        if dupes:
                            errors.append(
                                {
                                    "row": row_idx,
                                    "error": f"duplicate found: {dupes[0].customer_name}",
                                }
                            )
                            processed += 1
                            continue

                        customer = Customer.create(
                            name=name,
                            tenant_id=job.tenant_id,
                            created_by=job.user_id,
                            country=get_val("country"),
                            region=get_val("region"),
                            city=get_val("city"),
                            industry=get_val("industry"),
                            company_size=get_val("company_size"),
                            website=get_val("website"),
                            source=get_val("source"),
                            main_products=get_val("main_products"),
                            annual_volume=get_val("annual_volume"),
                            current_supplier=get_val("current_supplier"),
                            owner_id=job.user_id,
                        )
                        await customer_repo.create(customer)

                        contact_name = get_val("contact_name")
                        if contact_name or email:
                            contact = Contact(
                                customer_id=customer.id,
                                tenant_id=job.tenant_id,
                                name=contact_name or name,
                                email=email,
                                phone=get_val("contact_phone"),
                                whatsapp=get_val("whatsapp"),
                                is_primary=True,
                                created_by=job.user_id,
                            )
                            await contact_repo.create(contact)

                        success_count += 1
                        processed += 1

                    except Exception as e:  # noqa: BLE001
                        errors.append({"row": row_idx, "error": str(e)})
                        processed += 1

                await session.commit()
                job.processed_rows = processed
                job.success_count = success_count
                job.error_count = len(errors)
                await job_repo.update(job)
                await session.commit()

        except Exception as e:  # noqa: BLE001
            errors.append({"row": 0, "error": f"Fatal error: {e}"})

        job.status = DataJobStatus.COMPLETED
        job.completed_at = datetime.now(UTC)
        job.processed_rows = processed
        job.success_count = success_count
        job.error_count = len(errors)
        job.error_details = errors

        if errors:
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.append(["Row", "Error"])
            for err in errors:
                error_ws.append([err["row"], err["error"]])
            error_path = (job.file_path or "/tmp/import") + ".errors.xlsx"
            error_wb.save(error_path)
            job.result_file_url = error_path

        await job_repo.update(job)
        await session.commit()


async def process_customer_export(ctx: dict[str, Any], *, job_id: str) -> None:
    """Process customer export to Excel file."""
    async with async_session_factory() as session:
        job_repo = DataJobRepository(session)
        customer_repo = CustomerRepository(session)
        contact_repo = ContactRepository(session)
        grade_repo = CustomerGradeRepository(session)

        job = await job_repo.get_by_job_id(uuid.UUID(job_id))
        if job is None:
            return

        job.status = DataJobStatus.PROCESSING
        job.started_at = datetime.now(UTC)
        await job_repo.update(job)
        await session.commit()

        try:
            filters: CustomerFilter | None = None
            if job.filter_config:
                from app.domain.customer.enums import FollowUpStage

                fc = job.filter_config
                filters = CustomerFilter(
                    keyword=fc.get("keyword"),
                    grade_id=uuid.UUID(fc["grade_id"]) if fc.get("grade_id") else None,
                    source=fc.get("source"),
                    follow_status=FollowUpStage(fc["follow_status"]) if fc.get("follow_status") else None,
                    country=fc.get("country"),
                    industry=fc.get("industry"),
                    tag_ids=[uuid.UUID(t) for t in fc.get("tag_ids", [])],
                    owner_id=uuid.UUID(fc["owner_id"]) if fc.get("owner_id") else None,
                )

            total = await customer_repo.count_by_tenant(job.tenant_id)
            job.total_rows = total
            await job_repo.update(job)
            await session.commit()

            wb = Workbook()
            ws = wb.active
            ws.append(EXPORT_HEADERS)

            page = 1
            page_size = 500
            exported = 0

            while True:
                customers, _ = await customer_repo.get_by_tenant(job.tenant_id, page, page_size, filters=filters)
                if not customers:
                    break

                for customer in customers:
                    contacts = await contact_repo.get_by_customer_id(job.tenant_id, customer.id)
                    primary = next((c for c in contacts if c.is_primary), contacts[0] if contacts else None)

                    grade_name = ""
                    if customer.grade_id:
                        grade = await grade_repo.get_by_id(job.tenant_id, customer.grade_id)
                        if grade:
                            grade_name = grade.name

                    ws.append(
                        [
                            customer.name,
                            customer.country,
                            customer.region,
                            customer.city,
                            customer.industry,
                            customer.company_size,
                            customer.website,
                            customer.source,
                            grade_name,
                            customer.follow_status.value,
                            customer.main_products,
                            customer.annual_volume,
                            customer.current_supplier,
                            primary.name if primary else None,
                            primary.email if primary else None,
                            primary.phone if primary else None,
                            primary.whatsapp if primary else None,
                        ]
                    )
                    exported += 1

                job.processed_rows = exported
                await job_repo.update(job)
                await session.commit()
                page += 1

            export_path = tempfile.mktemp(suffix=".xlsx", prefix="customers_export_")
            wb.save(export_path)

            job.status = DataJobStatus.COMPLETED
            job.completed_at = datetime.now(UTC)
            job.success_count = exported
            job.result_file_url = export_path
            await job_repo.update(job)
            await session.commit()

        except Exception as e:  # noqa: BLE001
            job.status = DataJobStatus.FAILED
            job.completed_at = datetime.now(UTC)
            job.error_details = [{"error": str(e)}]
            await job_repo.update(job)
            await session.commit()
