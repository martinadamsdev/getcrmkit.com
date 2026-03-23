from __future__ import annotations

import contextlib
import tempfile
import uuid
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from app.domain.customer.enums import DataJobStatus
from app.domain.product.entities import Product
from app.infra.database.connection import async_session_factory
from app.infra.database.repositories.data_job_repository import DataJobRepository
from app.infra.database.repositories.product_category_repository import ProductCategoryRepository
from app.infra.database.repositories.product_repository import ProductRepository

BATCH_SIZE = 100

EXPORT_HEADERS = [
    "产品名称",
    "SKU",
    "分类",
    "描述",
    "材质",
    "尺寸",
    "重量(kg)",
    "颜色",
    "包装",
    "成本价",
    "成本货币",
    "售价",
    "售货币",
    "是否上架",
]

_FIELD_MAP: dict[str, str] = {
    "name": "name",
    "产品名称": "name",
    "sku": "sku",
    "description": "description",
    "描述": "description",
    "material": "material",
    "材质": "material",
    "dimensions": "dimensions",
    "尺寸": "dimensions",
    "weight": "weight",
    "重量(kg)": "weight",
    "color": "color",
    "颜色": "color",
    "packing": "packing",
    "包装": "packing",
    "cost_price": "cost_price",
    "成本价": "cost_price",
    "cost_currency": "cost_currency",
    "成本货币": "cost_currency",
    "selling_price": "selling_price",
    "售价": "selling_price",
    "selling_currency": "selling_currency",
    "售货币": "selling_currency",
}


async def process_product_import(ctx: dict[str, Any], *, job_id: str) -> None:
    """Process product import from Excel file."""
    async with async_session_factory() as session:
        job_repo = DataJobRepository(session)
        product_repo = ProductRepository(session)

        job = await job_repo.get_by_job_id(uuid.UUID(job_id))
        if job is None:
            return

        job.status = DataJobStatus.PROCESSING
        job.started_at = datetime.now(UTC)
        await job_repo.update(job)
        await session.commit()

        errors: list[dict[str, Any]] = []
        success_count = 0
        processed = 0

        try:
            wb = load_workbook(job.file_path, read_only=True)
            ws = wb.active
            header_row = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            col_indices: dict[str, int] = {}
            for i, h in enumerate(header_row):
                if h in _FIELD_MAP:
                    col_indices[_FIELD_MAP[h]] = i

            job.total_rows = len(rows)
            await job_repo.update(job)
            await session.commit()

            for batch_start in range(0, len(rows), BATCH_SIZE):
                batch = rows[batch_start : batch_start + BATCH_SIZE]
                for row_idx, row in enumerate(batch, start=batch_start + 2):
                    try:

                        def get_val(field: str, _row: Any = row) -> str | None:
                            idx = col_indices.get(field)
                            if idx is not None and idx < len(_row) and _row[idx]:
                                return str(_row[idx]).strip()
                            return None

                        name = get_val("name")
                        if not name:
                            errors.append({"row": row_idx, "error": "name is required"})
                            processed += 1
                            continue

                        weight_val: Decimal | None = None
                        weight_str = get_val("weight")
                        if weight_str:
                            with contextlib.suppress(InvalidOperation):
                                weight_val = Decimal(weight_str)

                        cost_price_val: Decimal | None = None
                        cost_str = get_val("cost_price")
                        if cost_str:
                            with contextlib.suppress(InvalidOperation):
                                cost_price_val = Decimal(cost_str)

                        selling_price_val: Decimal | None = None
                        sell_str = get_val("selling_price")
                        if sell_str:
                            with contextlib.suppress(InvalidOperation):
                                selling_price_val = Decimal(sell_str)

                        product = Product.create(
                            name=name,
                            tenant_id=job.tenant_id,
                            created_by=job.user_id,
                            sku=get_val("sku"),
                            description=get_val("description"),
                            material=get_val("material"),
                            dimensions=get_val("dimensions"),
                            weight=weight_val,
                            color=get_val("color"),
                            packing=get_val("packing"),
                            cost_price=cost_price_val,
                            cost_currency=get_val("cost_currency") or "CNY",
                            selling_price=selling_price_val,
                            selling_currency=get_val("selling_currency") or "USD",
                        )
                        await product_repo.create(product)
                        success_count += 1
                        processed += 1

                    except Exception as e:  # noqa: BLE001
                        errors.append({"row": row_idx, "error": str(e)})
                        processed += 1

                await session.commit()
                job.processed_rows = processed
                job.success_count = success_count
                job.error_count = len(errors)
                await job_repo.update(job)
                await session.commit()

        except Exception as e:  # noqa: BLE001
            errors.append({"row": 0, "error": f"Fatal error: {e}"})

        job.status = DataJobStatus.COMPLETED
        job.completed_at = datetime.now(UTC)
        job.processed_rows = processed
        job.success_count = success_count
        job.error_count = len(errors)
        job.error_details = errors

        if errors:
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.append(["Row", "Error"])
            for err in errors:
                error_ws.append([err["row"], err["error"]])
            error_path = (job.file_path or "/tmp/import") + ".errors.xlsx"
            error_wb.save(error_path)
            job.result_file_url = error_path

        await job_repo.update(job)
        await session.commit()


async def process_product_export(ctx: dict[str, Any], *, job_id: str) -> None:
    """Process product export to Excel file."""
    async with async_session_factory() as session:
        job_repo = DataJobRepository(session)
        product_repo = ProductRepository(session)
        category_repo = ProductCategoryRepository(session)

        job = await job_repo.get_by_job_id(uuid.UUID(job_id))
        if job is None:
            return

        job.status = DataJobStatus.PROCESSING
        job.started_at = datetime.now(UTC)
        await job_repo.update(job)
        await session.commit()

        try:
            from app.domain.product.value_objects import ProductFilter

            filters: ProductFilter | None = None
            if job.filter_config:
                fc = job.filter_config
                filters = ProductFilter(
                    keyword=fc.get("keyword"),
                    category_id=uuid.UUID(fc["category_id"]) if fc.get("category_id") else None,
                    is_active=fc.get("is_active"),
                )

            total = await product_repo.count_by_tenant(job.tenant_id)
            job.total_rows = total
            await job_repo.update(job)
            await session.commit()

            # Cache categories
            categories = await category_repo.get_all(job.tenant_id)
            cat_map = {c.id: c.name for c in categories}

            wb = Workbook()
            ws = wb.active
            ws.append(EXPORT_HEADERS)

            page = 1
            page_size = 500
            exported = 0

            while True:
                products, _ = await product_repo.get_by_tenant(job.tenant_id, page, page_size, filters=filters)
                if not products:
                    break

                for product in products:
                    ws.append(
                        [
                            product.name,
                            product.sku,
                            cat_map.get(product.category_id, "") if product.category_id else "",
                            product.description,
                            product.material,
                            product.dimensions,
                            str(product.weight) if product.weight else None,
                            product.color,
                            product.packing,
                            str(product.cost_price) if product.cost_price else None,
                            product.cost_currency,
                            str(product.selling_price) if product.selling_price else None,
                            product.selling_currency,
                            "Y" if product.is_active else "N",
                        ]
                    )
                    exported += 1

                job.processed_rows = exported
                await job_repo.update(job)
                await session.commit()
                page += 1

            export_path = tempfile.mktemp(suffix=".xlsx", prefix="products_export_")
            wb.save(export_path)

            job.status = DataJobStatus.COMPLETED
            job.completed_at = datetime.now(UTC)
            job.success_count = exported
            job.result_file_url = export_path
            await job_repo.update(job)
            await session.commit()

        except Exception as e:  # noqa: BLE001
            job.status = DataJobStatus.FAILED
            job.completed_at = datetime.now(UTC)
            job.error_details = [{"error": str(e)}]
            await job_repo.update(job)
            await session.commit()
